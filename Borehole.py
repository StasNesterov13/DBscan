import pandas as pd
from openpyxl import load_workbook
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler


class Borehole:

    def __init__(self):
        self.df = pd.DataFrame()
        self.name = None
        self.x_principal = None
        self.x_scaled = None
        self.pressure = None
        self.param = ['Состояние', 'Dшт, мм', 'Qж, м3/сут', 'Qгаз, м3/сут', 'Обв, %', 'Обв ХАЛ, %', 'Qн, т/сут',
                      'Рбуф, атм', 'Рзатр, атм', 'Рлин, атм', 'ГФ, м3/т', 'Рпл ГНК', 'Рзаб ГНК', 'F, Гц', 'Рприем, атм',
                      'Прим', 'Рзаб замер', 'Траб']

    def Creature(self, start, end, num_data, press):
        wb = load_workbook(filename='Debit.xlsx', read_only=True, data_only=True)
        ws = wb[f'{wb.sheetnames[0]}']
        date = []
        data = []
        self.pressure = press
        for row in ws[f'{start}2':f'{end}2']:
            for cell in row:
                date.append(cell.value)
        for i in [num_data + 2, num_data + 3, num_data + 4, num_data + 6,
                  num_data + 10, num_data + self.pressure]:
            data_row = []
            for row in ws[f'{start}{i}':f'{end}{i}']:
                for cell in row:
                    data_row.append(cell.value)
            data.append(data_row)
        self.name = ws[f'A{num_data}'].value
        self.df = pd.DataFrame(data, index=[self.param[2], self.param[3], self.param[4], self.param[6], self.param[10],
                                            self.param[self.pressure]], columns=date)
        self.df = self.df.transpose()
        self.df.dropna(axis=0, how='all', inplace=True)
        self.df.fillna(value=0, inplace=True)
        for cell in set(self.df.index.tolist()):
            if isinstance(cell, str):
                self.df.drop(index=cell, inplace=True)
        wb.close()

        with pd.ExcelWriter("One.xlsx", mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
            self.df.to_excel(writer, sheet_name=f'{self.name}')

        self.Standard()

    def Read(self, sheet):

        self.df = pd.read_excel('One.xlsx', sheet_name=f'{sheet}', index_col=0)
        self.name = sheet
        self.pressure = self.param.index(f'{self.df.columns[-1]}')
        self.Standard()

    def Standard(self):
        self.x_scaled = StandardScaler().fit_transform(self.df)
        self.x_scaled = pd.DataFrame(self.x_scaled)

        self.x_principal = PCA(n_components=2).fit_transform(self.x_scaled)
        self.x_principal = pd.DataFrame(self.x_principal)

        self.df['Cluster'] = 0
